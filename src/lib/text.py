# Лабораторная работа №3
import re
import csv
import json
from pathlib import Path


def normalize(text: str, casefold: bool = True, yo2e: bool = True) -> str:
    if not isinstance(text, str):
        raise TypeError("text должен быть строкой")

    if casefold:
        text = text.casefold()

    if yo2e:
        text = text.replace("ё", "е").replace("Ё", "Е")

    text = re.sub(r"[^\w\s]", " ", text)
    text = re.sub(r"\s+", " ", text)
    text = text.strip()

    return text


def tokenize(text: str) -> list[str]:
    # Проверка типа в tokenize
    if not isinstance(text, str):
        raise TypeError(f"Ожидается str, получен {type(text).__name__}")

    token = normalize(text)
    token = re.sub(r"[^\w\s]", "", token)
    token_1 = token.split()
    return token_1


def count_freq(tokens: list[str]) -> dict[str, int]:
    if not isinstance(tokens, list):
        raise TypeError(f"Ожидается list, получен {type(tokens).__name__}")

    count = {}
    for i in tokens:
        if not isinstance(i, str):
            raise TypeError(f"Все токены должны быть str, получен {type(i).__name__}")
        count[i] = count.get(i, 0) + 1
    return count


def top_n(freq: dict[str, int], n: int = 5) -> list[tuple[str, int]]:
    if not isinstance(freq, dict):
        raise TypeError(f"Ожидается dict, получен {type(freq).__name__}")
    if not isinstance(n, int):
        raise TypeError(f"n должен быть int, получен {type(n).__name__}")

    sorted_items = sorted(freq.items(), key=lambda x: (-x[1], x[0]))
    return sorted_items[:n]


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    import openpyxl  # type: ignore

    csv_file = Path(csv_path)
    if not csv_file.exists():
        raise FileNotFoundError(f"CSV файл не найден: {csv_path}")

    if csv_file.suffix.lower() != ".csv":
        raise ValueError(
            f"Неверный тип файла: ожидается .csv, получен {csv_file.suffix}"
        )

    xlsx_file = Path(xlsx_path)
    if xlsx_file.suffix.lower() != ".xlsx":
        raise ValueError(
            f"Неверный тип выходного файла: ожидается .xlsx, получен {xlsx_file.suffix}"
        )

    try:
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)
    except UnicodeDecodeError:
        raise ValueError("Ошибка кодировки: файл должен быть в UTF-8")
    except Exception as e:
        raise ValueError(f"Ошибка чтения CSV файла: {e}")

    if len(rows) == 0:
        raise ValueError("CSV файл пуст")

    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except ImportError:
        raise ImportError("openpyxl не установлен. Установите: pip install openpyxl")

    try:
        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise ValueError("Не удалось создать рабочий лист XLSX")
        ws.title = "Sheet1"

        for row_idx, row_data in enumerate(rows, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)

        for col_idx in range(1, len(rows[0]) + 1):
            column_letter = get_column_letter(col_idx)

            max_length = 8
            for row in ws[column_letter]:
                if row.value:
                    max_length = max(max_length, len(str(row.value)))

            ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(xlsx_path)

    except Exception as e:
        raise ValueError(f"Ошибка создания XLSX файла: {e}")


def json_to_csv(json_path: str, csv_path: str) -> None:
    json_file = Path(json_path)
    if not json_file.exists():
        raise FileNotFoundError(f"JSON файл не найден: {json_path}")

    if json_file.suffix.lower() != ".json":
        raise ValueError(
            f"Неверный тип файла: ожидается .json, получен {json_file.suffix}"
        )

    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except json.JSONDecodeError as e:
        raise ValueError(f"Ошибка парсинга JSON: {e}")

    if not isinstance(data, list):
        raise ValueError("JSON должен содержать список объектов")

    if len(data) == 0:
        raise ValueError("JSON файл пуст")

    if not all(isinstance(item, dict) for item in data):
        raise ValueError("Все элементы JSON должны быть словарями")

    all_fields = set()
    for item in data:
        all_fields.update(item.keys())

    first_fields = list(data[0].keys())
    remaining_fields = sorted(all_fields - set(first_fields))
    fieldnames = first_fields + remaining_fields

    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in data:
            normalized_row = {field: row.get(field, "") for field in fieldnames}
            writer.writerow(normalized_row)


def csv_to_json(csv_path: str, json_path: str) -> None:
    csv_file = Path(csv_path)
    if not csv_file.exists():
        raise FileNotFoundError(f"CSV файл не найден: {csv_path}")

    if csv_file.suffix.lower() != ".csv":
        raise ValueError(
            f"Неверный тип файла: ожидается .csv, получен {csv_file.suffix}"
        )

    try:
        with open(csv_path, "r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)

            if reader.fieldnames is None:
                raise ValueError("CSV файл не содержит заголовка")

            data = list(reader)
            if len(data) == 0:
                raise ValueError("CSV файл пуст (только заголовок)")

    except csv.Error as e:
        raise ValueError(f"Ошибка парсинга CSV: {e}")

    try:
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        raise ValueError(f"Ошибка записи JSON: {e}")
