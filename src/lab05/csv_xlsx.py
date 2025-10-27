import csv
from pathlib import Path
import sys
import os

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))



def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    csv_file = Path(csv_path)
    if not csv_file.exists():
        raise FileNotFoundError(f"CSV файл не найден: {csv_path}")
    
    if csv_file.suffix.lower() != '.csv':
        raise ValueError(f"Неверный тип файла: ожидается .csv, получен {csv_file.suffix}")
    
    xlsx_file = Path(xlsx_path)
    if xlsx_file.suffix.lower() != '.xlsx':
        raise ValueError(f"Неверный тип выходного файла: ожидается .xlsx, получен {xlsx_file.suffix}")
    

    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)
    except UnicodeDecodeError:
        raise ValueError("Ошибка кодировки: файл должен быть в UTF-8")
    except Exception as e:
        raise ValueError(f"Ошибка чтения CSV файла: {e}")
    

    if len(rows) == 0:
        raise ValueError("CSV файл пуст")
    

    """
    В данном блоке импортиуем необходимые модули из библиотеки Workbook
    также данный код from openpyxl.utils import get_column_letter нужен
    чтобы переделать номера колонки в буквы это нужно сделать так как 
    в excele вместо цифр в столбцах находятся буквы
    """
    try:
        from openpyxl import Workbook

        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl не установлен. Установите: pip install openpyxl")
    


    """
    wb = Workbook() благодаря данной строчки новую Excel книгу
    ws = wb.active создает 1 лист в книге и делаем его активным
    ws.title = "Sheet1" задаем название листу
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        """
        Первый цикл for row_idx, row_data in enumerate перебирает значение строки и индекс начиная с 1
        Второй цикл for col_idx, cell_value in enumerate перебирает ячейки в текущей строке начиная с 1
        После циклов записываем значения ws.cell(row, column, value) это метод для записи значений в конкретные ячейки
        row=row_idx номер строки, column=col_idx номер столбца, value=cell_value значение для записи
        """
        for row_idx, row_data in enumerate(rows, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        """
        Первый цикл переберает все столбцы после с помощью функции get_column_letter переводим все цифры в буквы
        Второй цикл переберает все ячейки в данной колонке дальше проверяем что в ячейке есть какое то значение 
        ищем максимальное значение 
        после этого задаем параметр ширины ячейки с помощью команды ws.column_dimensions[column_letter]
        """
        for col_idx in range(1, len(rows[0]) + 1):
            column_letter = get_column_letter(col_idx)

            max_length = 8  
            for row in ws[column_letter]:
                if row.value:
                    max_length = max(max_length, len(str(row.value)))
            
            ws.column_dimensions[column_letter].width = max_length + 2  

        """
        Сохраняем полученный файл
        """
        wb.save(xlsx_path)
        
    except Exception as e:
        raise ValueError(f"Ошибка создания XLSX файла: {e}")
    



csv_to_xlsx(r"python_labs\data\lab05\samples\Country.csv", r"python_labs\data\lab05\out\citiescsv.xlsx")
